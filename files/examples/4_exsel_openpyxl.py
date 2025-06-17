from openpyxl import Workbook, load_workbook

from files.files import EXCEL_FILE_PATH

# Чтение
wb = load_workbook(EXCEL_FILE_PATH)
sheet = wb.active

for row in sheet.iter_rows(values_only=True):
    print(row)

# Запись
wb = Workbook()
ws = wb.active
ws['A1'] = 'Test'
ws.append([1, 2, 3])
wb.save('output.xlsx')